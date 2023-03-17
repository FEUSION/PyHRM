########################################################################################################################
##Importing Libraries
import time
from datetime import datetime
import sys
import subprocess
import warnings
import os
warnings.filterwarnings('ignore')
try:
    import io
except:
    subprocess.check_call([sys.executable,'-m', 'pip', 'install', 'io', '-q'])
    import io
try:
    import tqdm
except:
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'tqdm','-q'])
    import tqdm
from tqdm import tqdm
for j in tqdm(range(10), desc=f'Setting Up..', leave=False):
    time.sleep(0.2)
subprocess.check_call([sys.executable,'-m', 'pip', 'install', '--upgrade', 'pip','-q'])
try:
    import packaging
except:
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'packaging','-q'])
try:
    import pandas as pd
except:
    subprocess.check_call([sys.executable, '-m', 'pip', 'install','pandas','-q'])
    import pandas as pd
try:
    import PIL
    from PIL import Image
    import requests
except:
    subprocess.check_call([sys.executable, '-m', 'pip', 'install','pillow','-q'])
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'requests','-q'])
    import PIL
    from PIL import Image
    import requests
try:
    import numpy as np
except:
    subprocess.check_call([sys.executable, '-m', 'pip', 'install','numpy','-q'])
    import numpy as np
try:
    import kaleido
except:
    subprocess.check_call([sys.executable,'-m','pip','install','-U','kaleido','-q'])
    import kaleido
try:
    import xlrd
except:
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'xlrd','-q'])
    import xlrd
try:
    import plotly
except:
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'plotly','-q'])
for j in tqdm(range(10), desc=f'Initializing Necessary Libraries..', leave=False):
    time.sleep(0.2)
try:
    import openpyxl
except:
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl','-q'])
    import openpyxl
try:
    import scipy
    from scipy.signal import find_peaks, peak_widths, peak_prominences
    from scipy.integrate import simpson
except:
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'scipy','-q'])
    import scipy
    from scipy.signal import find_peaks, peak_widths, peak_prominences
    from scipy.integrate import simpson
try:
    import tensorflow
    import keras
    from keras.models import load_model
except:
    print("Keras is a required component of this package \n Installing Keras..\n This a may take a while..")
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', '-U','tensorflow==2.12.0rc0'])
    subprocess.check_call([sys.executable, '-m','pip', 'install','-U','keras'])
    import tensorflow
    import keras
    from keras.models import load_model
try:
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
except:
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'matplotlib','-q'])
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
try:
    import fpdf
    from fpdf import FPDF
    import tempfile
except:
    subprocess.check_call([sys.executable, '-m','pip','install', 'FPDF','-q'])
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'tempfile', '-q'])
    import fpdf
    from fpdf import FPDF
    import tempfile
import plotly.graph_objects as go
import plotly.express as px
import plotly.io as pio
pio.renderers.default = 'browser'
for j in tqdm(range(5), desc=f'Setup Complete', leave=False):
    time.sleep(0.2)
########################################################################################################################


class MeltcurveInterpreter:

    def __init__(self):
        self.labels = []
        self.transformed_data = pd.DataFrame()
        self.model = load_model('Melt.h5', compile=False)
        for j in tqdm(range(2), desc=f'Initializing..', leave=False):
            time.sleep(0.2)

    def plot(self, data, save = False):

        from PIL import Image
        import requests
        from io import BytesIO
        response = requests.get("https://microlabindia.com/wp-content/uploads/2022/06/MBL-Logo.png")
        img = Image.open(BytesIO(response.content))

        fig = go.Figure()
        for X in range(1, len(data.columns)):
            fig.add_trace(go.Scatter(x=data.iloc[:, 0],
                                     y=data.iloc[:, X],
                                     name=self.labels[X-1]))
            if data.iloc[1,1]>20.0:
                title = "<i><b>Raw Fluorescence Curve</b></i>"
                ytitle = "Fluorescence"
                xtitle = 'Temperature in Celsius'
            elif data.iloc[0,0]==1:
                title = "<i><b>Amplification Curve</b></i>"
                ytitle = "Normalized Fluorescence"
                xtitle = 'Cycle Time'
            else:
                title = "<i><b>Melt Curve</b></i>"
                ytitle = 'dF/dT'
                xtitle = 'Temperature in Celsius'
            fig.update_layout(title_text=(title),
                              title_x = 0.5,
                              title_font_size=30,
                              title_font_family='Arial',
                              legend_itemclick="toggleothers",
                              legend_itemdoubleclick="toggleothers",
                              legend_groupclick="togglegroup",
                              legend_title_text='<b>Targets<b>',
                              legend_font_size=12,
                              legend_title_font_family='Arial',
                              legend_title_font_size=18,
                              legend_bgcolor="#f1f1f1",
                              legend_borderwidth=1,

                              plot_bgcolor='#ffffff',
                              title_font_color="#417a41",
                              )
            fig.update_xaxes(title_text =xtitle,
                             showgrid = False)
            fig.update_yaxes(title_text =ytitle,
                             showgrid= False)
            fig.layout.images = [dict(source=img,
                                      xref="paper",
                                      yref="paper",
                                      x=0.1,
                                      y=1.05,
                                      sizex=0.20,
                                      sizey=0.20,
                                      xanchor="center",
                                      yanchor="bottom")]
        if save:
            return fig
        fig.show()

    def save_path(self):
        path = input("Enter the path to save: ")
        return path

    def data_read(self, path, labels=False, index=False, figure=False):
        self.path = path
        from tqdm import tqdm
        for j in tqdm(range(10), desc=f'Loading data', leave=False):
            time.sleep(0.1)
        try:
            try:
                return_data = pd.read_excel(path, engine='xlrd')
            except:
                return_data = pd.read_excel(path)
        except:
            raise ValueError("Unsupported Format!")

        if index:
            return_data.drop(return_data.columns[0], axis =1, inplace= True)

        if len(return_data.columns)>3:
            if not all([return_data.columns[0] == 'Text', return_data.columns[1] == 'X', return_data.columns[2] == 'Y',
                        return_data.columns[3] == 'Text.1']):
                raise ValueError("""Could not Load:
                                    1. Please check your input spreadsheet format. 
                                        If your input file has index, please assign "index=True".
                                    2.Or Invalid File""")
        elif len(return_data.columns) == 3:
            if not all([return_data.columns[0] == 'Text', return_data.columns[1] == 'X', return_data.columns[2] == 'Y']):
                raise ValueError("""Could not Load:
                                    1. Please check your input spreadsheet format. 
                                        If your input file has index, please assign "index=True".
                                    2.Or Invalid File""")
        else:
            if not all([return_data.columns[0] == 'Text', return_data.columns[1] == 'X', return_data.columns[2] == 'Y',
                        return_data.columns[3] == 'Text.1']):
                raise ValueError("""Could not Load:
                                    1. Please check your input spreadsheet format. 
                                        If your input file has index, please assign "index=True".
                                    2.Or Invalid File""")

        # li_labels = []
        # for cols in return_data.iloc[:, 0::3].columns:
        #     li_labels.append(return_data[cols].unique()[0])
        # self.labels = li_labels

        li_labels = return_data.iloc[:,0::3].loc[1].apply(lambda x : str(x).split()[-1]).to_list()
        self.labels = li_labels

        dummy_data = pd.concat([return_data.iloc[:, 1], return_data.iloc[:, 2::3]], axis=1)
        del return_data
        return_data = dummy_data
        self.transformed_data = return_data

        if figure:
            self.plot(data=self.transformed_data)

        if labels:
            return [self.transformed_data, np.array(self.labels).reshape(-1)]
        else:
            return self.transformed_data

    def melt_convertion(self, figure = False, return_value = False, download=False):
        data_copy = self.transformed_data.copy()

        for columns in data_copy.columns[1:]:
            diff = np.gradient(data_copy[columns], data_copy.iloc[:,0])
            data_copy[columns] = -diff/10
        new_df = pd.DataFrame(columns = data_copy.columns)
        xnew = np.linspace(data_copy.iloc[0,0], data_copy.iloc[-1,0], int(len(data_copy.iloc[:,0])*2.945))
        new_df['X'] = xnew
        for cols in data_copy[1:]:
            splrep = scipy.interpolate.splrep(data_copy.iloc[:,0], data_copy[cols], s = 0.031)
            new_df[cols] = scipy.interpolate.splev(xnew,splrep)
        processed_data = new_df

        if figure:
            self.plot(data=processed_data)
        if download:
            saving_path = self.save_path()
            try:
                processed_data.to_csv(saving_path)
                print("Download Successful")
            except:
                print("Download Failed")

        if return_value:
            return processed_data

    def feature_detection(self,download=False, report = False):
        data = self.transformed_data
        c1 = ['Tm1', 'Tstart1', 'Tend1', 'Prom1', 'Width1', 'AUC1', 'Tm2', 'Tstart2', 'Tend2', 'Prom2',
              'Width2', 'AUC2', 'Target']
        max_peaks_of_all_curves = []
        features_data = pd.DataFrame(columns=c1)
        n = 3
        max_peaks_of_all_curves = []
        for k in range(1, len(data.columns)):
            x = data.iloc[:, 0].to_numpy()
            y = data.iloc[:, k].to_numpy()

            peaks, _ = find_peaks(y)
            # prominences
            proms = peak_prominences(y, peaks)[0]
            # sort the prominences in descending order
            desc_proms = np.argsort(proms)[::-1][:2]

            peaks = [peaks[u] for u in desc_proms]
            proms = [proms[l] for l in desc_proms]
            peak_tempeartures = [x[peak] for peak in peaks]

            width_data = np.array(peak_widths(y, peaks, rel_height=0.75)).T
            if len(width_data) > 0:
                first_highest_peak_att = width_data[0]
            if len(width_data) > 1:
                second_highest_peak_att = width_data[1]

            # third_highest_peak_att = width_data[2]

            def plot(x, y):
                fig, ax = plt.subplots()
                ax.plot(x, y)
                ax.set_xticks([])
                ax.set_yticks([])
                ax.set_axis_off()
                canvas = FigureCanvas(fig)
                png_output = io.BytesIO()
                canvas.print_png(png_output)
                png_output.seek(0)
                pil_image = Image.open(png_output)
                plt.close()
                plt.clf()
                return pil_image

            img = plot(x, y)
            img = img.resize((30, 30))
            img = np.array(img.convert("RGB"))
            img = img.reshape(1, 30, 30, 3)
            prediction = np.argmax(self.model.predict(img, verbose=0))

            if prediction == 0:
                features_data.loc[k, c1] = [0.0 for _ in range(len(c1))]
                features_data.loc[k, 'Target'] = self.labels[k-1]

            if prediction == 1:
                if len(first_highest_peak_att) == 0:
                    features_data.loc[k, c1] = [0.0 for _ in range(len(c1))]
                else:
                    width, prominence, start, end = *first_highest_peak_att,
                    start = x[round(start)]
                    end = x[round(end)]
                    features_data.loc[k, c1[:5]] = [peak_tempeartures[0], start, end, prominence, width]
                    features_data.loc[k, c1[5:]] = [0.0 for _ in range(len(c1[5:]))]

            if prediction == 2:
                if len(first_highest_peak_att) == 0:
                    features_data.loc[k, c1] = [0.0 for _ in range(len(c1))]

                else:
                    width, prominence, start, end = *first_highest_peak_att,
                    width1, prominence1, start1, end1 = *second_highest_peak_att,

                    if prominence1 < prominence * 0.25:
                        start = x[round(start)]
                        end = x[round(end)]
                        features_data.loc[k, c1[:5]] = [peak_tempeartures[0], start, end, prominence, width]
                        features_data.loc[k, c1[5:]] = [0.0 for _ in range(len(c1[5:]))]

                    else:
                        start = x[round(start)]
                        end = x[round(end)]
                        start1 = x[round(start1)]
                        end1 = x[round(end1)]
                        features_data.loc[k, c1[:5]] = [peak_tempeartures[0], start, end, prominence, width]
                        features_data.loc[k, c1[6:11]] = [peak_tempeartures[1], start1, end1, prominence1, width1]
                        features_data.loc[k, c1[11:]] = [0.0 for _ in range(len(c1[11:]))]

            temp_range = data.iloc[:, 0]

            def aidsimpson(temp_range, k, features_data,column1, column2):
                Tstartindex = temp_range[temp_range == features_data.loc[k, column1]].index[0]
                Tendindex = temp_range[temp_range == features_data.loc[k, column2]].index[0]
                return [Tstartindex, Tendindex]

            columns1 = ['Tstart1', 'Tstart2', 'Tstart3']
            columns2 = ['Tend1', 'Tend2', 'Tend3']

            for ite in range(1, 3):
                try:
                    indexes = aidsimpson(temp_range, k, features_data, column1=columns1[ite - 1],
                                         column2=columns2[ite - 1])
                    features_data.loc[k, 'AUC' + str(ite)] = simpson(data.iloc[indexes[0]:indexes[1], k].to_numpy(),
                                                                     data.iloc[indexes[0]:indexes[1], 0].to_numpy())
                except:
                    features_data.loc[k, 'AUC' + str(ite)] = 0.0
            features_data.loc[k,'Target'] = self.labels[k - 1]

        maximum = np.sort(features_data['Prom1'].to_numpy())[-1]
        for i in range(1, features_data.shape[0] + 1):
            if features_data.loc[i, 'Prom1'] < maximum * 0.04:
                features_data.loc[i, features_data.columns] = [0.0 for _ in range(13)]
                features_data.loc[i, 'Target'] = self.labels[i - 1]
        if download:
            print("Only supports .csv format for now")
            saving_path = self.save_path()
            try:
                features_data.to_csv(saving_path)
                print("Download Successful")
            except:
                print("Download Failed")

        if report:
            dataa = features_data.copy()
            for cols in dataa.columns[:-1]:
                dataa[cols] = dataa[cols].apply(lambda x: round(x, 2))
            graph = self.plot(self.transformed_data, save=True)
            with tempfile.NamedTemporaryFile(delete=False) as f:
                graph.write_image(f.name, format='png', width=1000)
                temp_image_file = f.name

            class PDF(FPDF):
                def __init__(self):
                    super().__init__()

                def header(self):
                    self.set_font('Arial', '', 12)
                    self.cell(0, 0, '', 0, 1, 'C')

                def footer(self):
                    self.set_y(-15)
                    self.set_font('Arial', '', 12)

            pdf = PDF()
            pdf.add_page()
            pdf.set_font('Arial', 'B', 24)
            pdf.cell(w=0, h=15, txt="Melt Signal Processing", ln=1)
            pdf.ln(2)
            pdf.set_font('Arial', '', 10)
            pdf.cell(w=30, h=5, txt="Date: ", ln=0)
            pdf.cell(w=30, h=5, txt=str(datetime.now().strftime("%d/%m/%Y")), ln=1)
            pdf.cell(w=30, h=5, txt="File: ", ln=0)
            pdf.cell(w=30, h=5, txt=str(self.path.split('\\')[-1]), ln=1)
            pdf.ln(5)
            pdf.image(temp_image_file, x=1, y=None, w=200, h=95, type='PNG', link='')

            # Table contents
            pdf.set_font('Arial', '', 9)
            for cols in dataa.columns:
                pdf.cell(15, 10, cols, 1)
            pdf.ln(1.8)
            for index, row in dataa.iterrows():
                pdf.ln(8)
                pdf.cell(15, 8, str(row['Tm1']), 1)
                pdf.cell(15, 8, str(row['Tstart1']), 1)
                pdf.cell(15, 8, str(row['Tend1']), 1)
                pdf.cell(15, 8, str(row['Prom1']), 1)
                pdf.cell(15, 8, str(row['Width1']), 1)
                pdf.cell(15, 8, str(row['AUC1']), 1)
                pdf.cell(15, 8, str(row['Tm2']), 1)
                pdf.cell(15, 8, str(row['Tstart2']), 1)
                pdf.cell(15, 8, str(row['Tend2']), 1)
                pdf.cell(15, 8, str(row['Prom2']), 1)
                pdf.cell(15, 8, str(row['Width2']), 1)
                pdf.cell(15, 8, str(row['AUC2']), 1)
                pdf.cell(15, 8, str(row['Target']), 1)
            saving_path2 = self.save_path()
            pdf.output(saving_path2, 'F')
            os.remove(temp_image_file)


        return features_data

